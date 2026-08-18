import time
import requests
import concurrent.futures
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE_URL = "http://localhost:8000"
NUM_USERS = 100
DURATION_SECONDS = 60
OUTPUT_EXCEL = r"c:\Users\ASUA\Desktop\NagiReddy_App\meal-resq-backend\load-tests\load-test-results.xlsx"



endpoints = [
    ("/api/v1/auth/login", "POST", {"login_id": "support.mealresq@gmail.com", "password": "password123", "role": "donor"}),
    ("/api/v1/ngo/available-donations", "GET", None),
    ("/api/v1/donor/donations", "GET", None),
]

response_times = []
status_codes = []
start_time = time.time()

def make_request(ep_tuple):
    url_path, method, payload = ep_tuple
    t0 = time.time()
    try:
        if method == "POST":
            res = requests.post(f"{BASE_URL}{url_path}", json=payload, timeout=5)
        else:
            res = requests.get(f"{BASE_URL}{url_path}", timeout=5)
        t1 = time.time()
        elapsed_ms = (t1 - t0) * 1000
        return elapsed_ms, res.status_code
    except Exception as e:
        t1 = time.time()
        return (t1 - t0) * 1000, 500

print(f"Starting Baseline Load Test: {NUM_USERS} Virtual Users for {DURATION_SECONDS} seconds...")

total_requests = 0
with concurrent.futures.ThreadPoolExecutor(max_workers=NUM_USERS) as executor:
    end_time = time.time() + DURATION_SECONDS
    while time.time() < end_time:
        futures = [executor.submit(make_request, endpoints[i % len(endpoints)]) for i in range(NUM_USERS)]
        for f in concurrent.futures.as_completed(futures):
            res_ms, code = f.result()
            response_times.append(res_ms)
            status_codes.append(code)
            total_requests += 1

test_duration = time.time() - start_time
rps = round(total_requests / test_duration, 2) if test_duration > 0 else 0
avg_ms = round(sum(response_times) / len(response_times), 2) if response_times else 0
min_ms = round(min(response_times), 2) if response_times else 0
max_ms = round(max(response_times), 2) if response_times else 0
pass_count = sum(1 for c in status_codes if c < 500)
fail_count = total_requests - pass_count

print("\nBaseline Load Test Execution Summary:")

print(f" • Total Virtual Users: {NUM_USERS}")
print(f" • Duration: {round(test_duration, 2)}s")
print(f" • Total Requests: {total_requests}")
print(f" • Requests Per Second (RPS): {rps} req/sec")
print(f" • Average Response Time: {avg_ms} ms")
print(f" • Min Response Time: {min_ms} ms")
print(f" • Max Response Time: {max_ms} ms")
print(f" • Success Rate: {round((pass_count/total_requests)*100, 2)}%\n")

# Excel Report Generation
wb = openpyxl.Workbook()
ws_sum = wb.active
ws_sum.title = "Load Test Summary"

header_fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
center_align = Alignment(horizontal="center", vertical="center")
left_align = Alignment(horizontal="left", vertical="center")

thin_border = Border(
    left=Side(style='thin', color='D1D5DB'),
    right=Side(style='thin', color='D1D5DB'),
    top=Side(style='thin', color='D1D5DB'),
    bottom=Side(style='thin', color='D1D5DB')
)

ws_sum.append(["Metric Category", "Measured Value", "Status / SLA Target"])
sum_rows = [
    ["Concurrent Virtual Users", NUM_USERS, "100 Users Target"],
    ["Test Duration", f"{round(test_duration, 2)} seconds", "60s Continuous"],
    ["Total Requests Processed", total_requests, "Thousands Processed"],
    ["Requests Per Second (RPS)", f"{rps} req/sec", "Target > 100 RPS"],
    ["Average Latency", f"{avg_ms} ms", "SLA < 300 ms"],
    ["Min Latency", f"{min_ms} ms", "Fastest Response"],
    ["Max Latency", f"{max_ms} ms", "Peak Spike"],
    ["Passed Requests", pass_count, "200 OK"],
    ["Failed Requests", fail_count, "0 Errors Allowed"],
    ["Overall Success Rate", f"{round((pass_count/total_requests)*100, 2)}%", "100.00% Pass Rate"],
]

for row in sum_rows:
    ws_sum.append(row)

for col in range(1, 4):
    cell = ws_sum.cell(row=1, column=col)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center_align

for row in range(2, len(sum_rows) + 2):
    for col in range(1, 4):
        cell = ws_sum.cell(row=row, column=col)
        cell.border = thin_border
        cell.alignment = left_align if col == 1 else center_align

for col in ws_sum.columns:
    max_len = max(len(str(cell.value or '')) for cell in col)
    col_letter = get_column_letter(col[0].column)
    ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 15)

wb.save(OUTPUT_EXCEL)
print(f"Excel Load Test Report generated at: {OUTPUT_EXCEL}")
